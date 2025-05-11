from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QTextEdit, QFileDialog, QMessageBox, QSplitter,
    QDialog, QLineEdit, QFormLayout, QInputDialog
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor, QTextCharFormat
import os
import shutil
import hashlib
import smtplib
from email.mime.text import MIMEText
from datetime import datetime
import pandas as pd

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("产品数据比对系统 3.0")
        self.setGeometry(100, 100, 1200, 800)
        self.lock_time = None
        self.failed_attempts = 0
        self.max_attempts = 3
        self.lock_duration = 15 * 60  # 15分钟锁定
        self.user_email = None
        self.user_password_hash = None
        self.admin_password_hash = "8c6976e5b5410415bde908bd4dee15dfb167a9c873fc4bb8a81f6f2ab448a918"  # admin默认密码
        self.manual_files = []
        self.db_compare_file = None
        self.current_db_file = None
        
        # 确保数据目录存在
        os.makedirs('data', exist_ok=True)
        
        # 加载保存的凭证
        self.load_credentials()
        
        # 主界面布局
        self.setup_ui()
        
        # 检查是否首次运行
        self.check_first_run()
        
        # 自动加载最新的数据库文件
        self.load_latest_database()

    def setup_ui(self):
        # 主部件和布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # 分割左右区域
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # 左侧功能区域
        left_widget = QWidget()
        left_layout = QVBoxLayout()
        
        # 1. 数据库上传区域
        self.setup_database_upload_area(left_layout)
        
        # 2. 手动比对区域
        self.setup_manual_comparison_area(left_layout)
        
        # 3. 数据库比对区域
        self.setup_db_comparison_area(left_layout)
        
        # 4. 账户管理区域
        self.setup_account_management_area(left_layout)
        
        left_widget.setLayout(left_layout)
        
        # 右侧信息显示区域
        self.info_display = QTextEdit()
        self.info_display.setReadOnly(True)
        
        splitter.addWidget(left_widget)
        splitter.addWidget(self.info_display)
        splitter.setSizes([400, 800])
        
        main_layout = QHBoxLayout()
        main_layout.addWidget(splitter)
        main_widget.setLayout(main_layout)

    def check_first_run(self):
        """检查是否是首次运行，需要设置密码和邮箱"""
        if not os.path.exists('data/credentials.txt'):
            self.show_setup_dialog()
            # 如果用户没有完成设置，阻止进入主界面
            if not os.path.exists('data/credentials.txt'):
                QMessageBox.critical(self, "错误", "必须设置密码和邮箱才能使用本软件")
                sys.exit(1)

    def show_setup_dialog(self):
        """显示初始设置对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("初始设置")
        layout = QFormLayout()
        
        email_edit = QLineEdit()
        email_edit.setPlaceholderText("请输入您的邮箱")
        password_edit = QLineEdit()
        password_edit.setPlaceholderText("请输入密码")
        password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        confirm_edit = QLineEdit()
        confirm_edit.setPlaceholderText("请确认密码")
        confirm_edit.setEchoMode(QLineEdit.EchoMode.Password)
        
        submit_btn = QPushButton("提交")
        
        def save_credentials():
            email = email_edit.text()
            password = password_edit.text()
            confirm = confirm_edit.text()
            
            if not email or not password:
                QMessageBox.warning(dialog, "错误", "邮箱和密码不能为空")
                return
                
            # 验证邮箱格式
            if '@' not in email or '.' not in email:
                QMessageBox.warning(dialog, "错误", "请输入有效的邮箱地址")
                return
                
            # 验证密码长度
            if len(password) < 6:
                QMessageBox.warning(dialog, "错误", "密码长度不能少于6位")
                return
                
            if password != confirm:
                QMessageBox.warning(dialog, "错误", "两次输入的密码不一致")
                return
                
            # 保存凭证
            os.makedirs('data', exist_ok=True)
            with open('data/credentials.txt', 'w') as f:
                f.write(f"{email}\n{hashlib.sha256(password.encode()).hexdigest()}")
                
            self.user_email = email
            self.user_password_hash = hashlib.sha256(password.encode()).hexdigest()
            dialog.accept()
        
        submit_btn.clicked.connect(save_credentials)
        
        layout.addRow("邮箱:", email_edit)
        layout.addRow("密码:", password_edit)
        layout.addRow("确认密码:", confirm_edit)
        layout.addRow(submit_btn)
        
        dialog.setLayout(layout)
        dialog.exec()

    def verify_password(self):
        """验证密码"""
        if self.is_locked():
            QMessageBox.warning(self, "锁定", f"系统已锁定，请等待15分钟或重启程序")
            return False
            
        password, ok = QInputDialog.getText(
            self, '密码验证', '请输入上传密码:', 
            QLineEdit.EchoMode.Password
        )
        
        if not ok:
            return False
            
        if hashlib.sha256(password.encode()).hexdigest() != self.user_password_hash:
            self.failed_attempts += 1
            remaining_attempts = self.max_attempts - self.failed_attempts
            
            if remaining_attempts > 0:
                QMessageBox.warning(self, "密码错误", 
                    f"密码错误，还剩{remaining_attempts}次尝试机会")
            else:
                self.lock_time = datetime.now()
                self.send_lock_notification()
                QMessageBox.warning(self, "锁定", 
                    "密码错误次数过多，系统已锁定15分钟\n锁定期间重启仍需验证密码")
                
            return False
            
        self.failed_attempts = 0
        return True

    def is_locked(self):
        """检查是否在锁定状态"""
        if self.lock_time and (datetime.now() - self.lock_time).total_seconds() < self.lock_duration:
            return True
        elif self.lock_time:
            self.lock_time = None
            self.failed_attempts = 0
        return False

    def send_lock_notification(self):
        """发送锁定通知邮件"""
        if not self.user_email:
            return
            
        try:
            msg = MIMEText("您的数据比对系统因多次密码错误已被锁定15分钟")
            msg['Subject'] = '系统锁定通知'
            msg['From'] = 'system@datacompare.com'
            msg['To'] = self.user_email
            
            # 这里需要配置SMTP服务器
            with smtplib.SMTP('smtp.example.com', 587) as server:
                server.starttls()
                server.login('username', 'password')
                server.send_message(msg)
        except Exception as e:
            self.log_message(f"发送邮件失败: {str(e)}")

    def setup_database_upload_area(self, layout):
        """设置数据库上传区域"""
        group = QWidget()
        vbox = QVBoxLayout()
        
        label = QLabel("数据库上传区域")
        upload_btn = QPushButton("上传数据库文件")
        upload_btn.clicked.connect(self.upload_database_file)
        
        vbox.addWidget(label)
        vbox.addWidget(upload_btn)
        group.setLayout(vbox)
        layout.addWidget(group)

    def setup_manual_comparison_area(self, layout):
        """设置手动比对区域"""
        group = QWidget()
        vbox = QVBoxLayout()
        
        label = QLabel("手动比对区域")
        upload_btn1 = QPushButton("上传比对文件1")
        upload_btn2 = QPushButton("上传比对文件2") 
        upload_btn3 = QPushButton("上传比对文件3")
        compare_btn = QPushButton("开始比对")
        
        upload_btn1.clicked.connect(lambda: self.upload_manual_file(1))
        upload_btn2.clicked.connect(lambda: self.upload_manual_file(2))
        upload_btn3.clicked.connect(lambda: self.upload_manual_file(3))
        compare_btn.clicked.connect(self.compare_manual_files)
        
        vbox.addWidget(label)
        vbox.addWidget(upload_btn1)
        vbox.addWidget(upload_btn2)
        vbox.addWidget(upload_btn3)
        vbox.addWidget(compare_btn)
        group.setLayout(vbox)
        layout.addWidget(group)

    def setup_db_comparison_area(self, layout):
        """设置数据库比对区域"""
        group = QWidget()
        vbox = QVBoxLayout()
        
        label = QLabel("数据库比对区域")
        upload_btn = QPushButton("上传比对文件")
        compare_btn = QPushButton("与数据库比对")
        
        upload_btn.clicked.connect(self.upload_compare_file)
        compare_btn.clicked.connect(self.compare_with_database)
        
        vbox.addWidget(label)
        vbox.addWidget(upload_btn)
        vbox.addWidget(compare_btn)
        group.setLayout(vbox)
        layout.addWidget(group)

    def setup_account_management_area(self, layout):
        """设置账户管理区域"""
        group = QWidget()
        vbox = QVBoxLayout()
        
        label = QLabel("账户管理")
        change_pwd_btn = QPushButton("修改密码")
        reset_pwd_btn = QPushButton("忘记密码?点击这里重置")
        change_email_btn = QPushButton("更换邮箱")
        admin_btn = QPushButton("管理员设置")
        
        # 添加使用说明提示
        help_label = QLabel("密码找回说明:\n1. 点击'忘记密码'按钮\n2. 输入注册邮箱收到的验证码\n3. 设置新密码(至少6位)")
        help_label.setStyleSheet("color: #666; font-size: 10px;")
        
        change_pwd_btn.clicked.connect(self.change_password)
        reset_pwd_btn.clicked.connect(self.reset_password)
        change_email_btn.clicked.connect(self.change_email)
        admin_btn.clicked.connect(self.admin_settings)
        
        vbox.addWidget(label)
        vbox.addWidget(change_pwd_btn)
        vbox.addWidget(reset_pwd_btn)
        vbox.addWidget(change_email_btn)
        vbox.addWidget(admin_btn)
        group.setLayout(vbox)
        layout.addWidget(group)

    def change_password(self):
        """修改密码"""
        if not self.verify_password():
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("修改密码")
        layout = QFormLayout()
        
        old_pwd = QLineEdit()
        old_pwd.setPlaceholderText("请输入旧密码")
        old_pwd.setEchoMode(QLineEdit.EchoMode.Password)
        new_pwd = QLineEdit()
        new_pwd.setPlaceholderText("请输入新密码")
        new_pwd.setEchoMode(QLineEdit.EchoMode.Password)
        confirm_pwd = QLineEdit()
        confirm_pwd.setPlaceholderText("请确认新密码")
        confirm_pwd.setEchoMode(QLineEdit.EchoMode.Password)
        
        submit_btn = QPushButton("提交")
        
        def save_new_password():
            if (hashlib.sha256(old_pwd.text().encode()).hexdigest() != 
                self.user_password_hash):
                QMessageBox.warning(dialog, "错误", "旧密码不正确")
                return
                
            if len(new_pwd.text()) < 6:
                QMessageBox.warning(dialog, "错误", "密码长度不能少于6位")
                return
                
            if new_pwd.text() != confirm_pwd.text():
                QMessageBox.warning(dialog, "错误", "两次输入的新密码不一致")
                return
                
            # 更新密码
            self.user_password_hash = hashlib.sha256(new_pwd.text().encode()).hexdigest()
            with open('data/credentials.txt', 'w') as f:
                f.write(f"{self.user_email}\n{self.user_password_hash}")
                
            QMessageBox.information(dialog, "成功", "密码修改成功")
            dialog.accept()
        
        submit_btn.clicked.connect(save_new_password)
        
        layout.addRow("旧密码:", old_pwd)
        layout.addRow("新密码:", new_pwd)
        layout.addRow("确认新密码:", confirm_pwd)
        layout.addRow(submit_btn)
        
        dialog.setLayout(layout)
        dialog.exec()

    def reset_password(self):
        """重置密码"""
        # 验证邮箱是否设置
        if not self.user_email:
            QMessageBox.warning(self, "错误", "未设置邮箱，无法找回密码")
            return
            
        # 发送包含临时验证码的邮件
        try:
            temp_code = hashlib.sha256(os.urandom(32)).hexdigest()[:8]
            msg = MIMEText(f"您的密码重置验证码是: {temp_code}\n验证码15分钟内有效")
            msg['Subject'] = '密码重置验证码'
            msg['From'] = 'system@datacompare.com'
            msg['To'] = self.user_email
            
            with smtplib.SMTP('smtp.example.com', 587) as server:
                server.starttls()
                server.login('username', 'password')
                server.send_message(msg)
                
            # 弹出验证码输入框
            code, ok = QInputDialog.getText(
                self, '验证码验证', 
                f"验证码已发送到{self.user_email}，请输入验证码:"
            )
            
            if ok and code == temp_code:
                # 允许设置新密码
                new_pwd, ok = QInputDialog.getText(
                    self, '设置新密码',
                    '请输入新密码(至少6位):',
                    QLineEdit.EchoMode.Password
                )
                if ok and len(new_pwd) >= 6:
                    self.user_password_hash = hashlib.sha256(new_pwd.encode()).hexdigest()
                    with open('data/credentials.txt', 'w') as f:
                        f.write(f"{self.user_email}\n{self.user_password_hash}")
                    QMessageBox.information(self, "成功", "密码重置成功")
                else:
                    QMessageBox.warning(self, "错误", "密码长度不足6位")
            else:
                QMessageBox.warning(self, "错误", "验证码不正确")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"发送邮件失败: {str(e)}")

    def change_email(self):
        """更换邮箱"""
        if not self.verify_password():
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("更换邮箱")
        layout = QFormLayout()
        
        new_email = QLineEdit()
        new_email.setPlaceholderText("请输入新邮箱")
        password = QLineEdit()
        password.setPlaceholderText("请输入密码验证")
        password.setEchoMode(QLineEdit.EchoMode.Password)
        
        submit_btn = QPushButton("提交")
        
        def save_new_email():
            if hashlib.sha256(password.text().encode()).hexdigest() != self.user_password_hash:
                QMessageBox.warning(dialog, "错误", "密码不正确")
                return
                
            if '@' not in new_email.text() or '.' not in new_email.text():
                QMessageBox.warning(dialog, "错误", "请输入有效的邮箱地址")
                return
                
            # 更新邮箱
            self.user_email = new_email.text()
            with open('data/credentials.txt', 'w') as f:
                f.write(f"{self.user_email}\n{self.user_password_hash}")
                
            QMessageBox.information(dialog, "成功", "邮箱修改成功")
            dialog.accept()
        
        submit_btn.clicked.connect(save_new_email)
        
        layout.addRow("新邮箱:", new_email)
        layout.addRow("密码验证:", password)
        layout.addRow(submit_btn)
        
        dialog.setLayout(layout)
        dialog.exec()

    def admin_settings(self):
        """管理员设置"""
        password, ok = QInputDialog.getText(
            self, '管理员验证', 
            '请输入管理员密码:',
            QLineEdit.EchoMode.Password
        )
        
        if ok and hashlib.sha256(password.encode()).hexdigest() == self.admin_password_hash:
            dialog = QDialog(self)
            dialog.setWindowTitle("管理员设置")
            layout = QVBoxLayout()
            
            reset_btn = QPushButton("重置用户密码和邮箱")
            change_admin_btn = QPushButton("修改管理员密码")
            
            def reset_user_credentials():
                os.remove('data/credentials.txt')
                self.show_setup_dialog()
                dialog.accept()
                
            def change_admin_password():
                new_pwd, ok = QInputDialog.getText(
                    self, '修改管理员密码',
                    '请输入新密码(至少6位):',
                    QLineEdit.EchoMode.Password
                )
                if ok and len(new_pwd) >= 6:
                    self.admin_password_hash = hashlib.sha256(new_pwd.encode()).hexdigest()
                    # 保存管理员密码到配置文件
                    with open('data/admin_credentials.txt', 'w') as f:
                        f.write(self.admin_password_hash)
                    QMessageBox.information(self, "成功", "管理员密码修改成功")
                else:
                    QMessageBox.warning(self, "错误", "密码长度不足6位")
            
            reset_btn.clicked.connect(reset_user_credentials)
            change_admin_btn.clicked.connect(change_admin_password)
            
            layout.addWidget(reset_btn)
            layout.addWidget(change_admin_btn)
            dialog.setLayout(layout)
            dialog.exec()
        else:
            QMessageBox.warning(self, "错误", "管理员密码不正确")

    def load_latest_database(self):
        """自动加载最新的数据库文件"""
        try:
            # 检查xlsx目录
            xlsx_dir = os.path.join('data', 'xlsx')
            if os.path.exists(xlsx_dir):
                # 只获取非临时文件(~$开头的)并按文件名中的时间戳排序
                files = [f for f in os.listdir(xlsx_dir) 
                        if f.endswith('.xlsx') and not f.startswith('~$')]
                if files:
                    # 按文件名中的时间戳排序(格式: YYYYMMDD_HHMMSS.xlsx)
                    files.sort(key=lambda x: x.split('.')[0], reverse=True)
                    latest_file = os.path.join(xlsx_dir, files[0])
                    self.current_db_file = latest_file
                    self.log_message(f"已自动加载数据库文件: {latest_file}")
                    self.display_file_info(latest_file)
                    return True
            
            self.log_message("未找到数据库文件，请先上传数据库文件", "orange")
            return False
        except Exception as e:
            self.log_message(f"加载数据库文件失败: {str(e)}", "red")
            return False

    def load_credentials(self):
        """加载保存的凭证"""
        try:
            # 加载用户凭证
            if os.path.exists('data/credentials.txt'):
                with open('data/credentials.txt', 'r') as f:
                    lines = f.readlines()
                    self.user_email = lines[0].strip()
                    self.user_password_hash = lines[1].strip()
            
            # 加载管理员凭证
            if os.path.exists('data/admin_credentials.txt'):
                with open('data/admin_credentials.txt', 'r') as f:
                    self.admin_password_hash = f.read().strip()
        except Exception as e:
            self.log_message(f"加载凭证失败: {str(e)}", "red")

    def log_message(self, message, color=None):
        """记录带颜色的消息"""
        if color:
            cursor = self.info_display.textCursor()
            format = QTextCharFormat()
            format.setForeground(QColor(color))
            cursor.setCharFormat(format)
            cursor.insertText(message + "\n")
        else:
            self.info_display.append(message)

    def display_file_info(self, file_path):
        """显示文件信息"""
        try:
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path) / 1024  # KB
            ext = os.path.splitext(file_path)[1].lower()
            
            info = f"文件信息:\n名称: {file_name}\n大小: {file_size:.2f}KB\n类型: {ext}"
            
            if ext in ('.xlsx', '.csv'):
                if ext == '.xlsx':
                    df = pd.read_excel(file_path)
                else:
                    df = pd.read_csv(file_path)
                info += f"\n行数: {len(df)}\n列数: {len(df.columns)}"
                
            self.log_message(info)
        except Exception as e:
            self.log_message(f"获取文件信息失败: {str(e)}", "red")

    def upload_database_file(self):
        """上传数据库文件"""
        try:
            if not self.verify_password():
                return
                
            file_path, _ = QFileDialog.getOpenFileName(
                self, "选择数据库文件", 
                "", "Excel Files (*.xlsx);;CSV Files (*.csv)"
            )
            
            if file_path:
                # 获取文件扩展名
                ext = os.path.splitext(file_path)[1][1:].lower()
                save_dir = os.path.join('data', ext)
                
                # 创建目录
                os.makedirs(save_dir, exist_ok=True)
                
                # 备份文件
                backup_dir = os.path.join('data', 'backup')
                os.makedirs(backup_dir, exist_ok=True)
                backup_path = os.path.join(backup_dir, os.path.basename(file_path))
                
                # 保存文件
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                save_path = os.path.join(save_dir, f"{timestamp}.{ext}")
                
                # 执行复制
                shutil.copy2(file_path, backup_path)
                shutil.copy2(file_path, save_path)
                
                # 清理旧文件
                for f in os.listdir(save_dir):
                    if f != os.path.basename(save_path):
                        try:
                            os.remove(os.path.join(save_dir, f))
                        except Exception as e:
                            self.log_message(f"删除旧文件失败: {str(e)}", "red")
                
                self.log_message(f"数据库文件已上传并保存: {save_path}")
                self.display_file_info(file_path)
                QMessageBox.information(self, "成功", "数据库文件上传成功")
        except Exception as e:
            self.log_message(f"上传数据库文件出错: {str(e)}", "red")
            QMessageBox.critical(self, "错误", f"上传失败: {str(e)}")

    def upload_manual_file(self, file_num):
        """上传手动比对文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, f"选择比对文件{file_num}",
            "", "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        
        if file_path:
            if len(self.manual_files) >= file_num:
                self.manual_files[file_num-1] = file_path
            else:
                self.manual_files.append(file_path)
            self.log_message(f"比对文件{file_num}已上传: {file_path}")
            self.display_file_info(file_path)

    def upload_compare_file(self):
        """上传比对文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择比对文件",
            "", "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        
        if file_path:
            self.db_compare_file = file_path
            self.log_message(f"比对文件已上传: {file_path}")
            self.display_file_info(file_path)

    def compare_manual_files(self):
        """手动比对文件"""
        if len(self.manual_files) < 2:
            QMessageBox.warning(self, "错误", "至少需要上传2个文件才能比对")
            return
            
        try:
            # 读取所有文件
            dfs = []
            file_names = []
            for file in self.manual_files:
                if file.endswith('.xlsx'):
                    df = pd.read_excel(file)
                else:
                    df = pd.read_csv(file)
                dfs.append(df)
                file_names.append(os.path.basename(file))
                
            # 检查格式一致性
            cols = [df.columns.tolist() for df in dfs]
            if not all(c == cols[0] for c in cols):
                QMessageBox.warning(self, "错误", "文件格式不一致，无法比对")
                return
                
            # 创建结果目录
            result_dir = os.path.join('results', 'compare_reports')
            os.makedirs(result_dir, exist_ok=True)
            
            # 生成JSON报告文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_path = os.path.join(result_dir, f"manual_compare_{timestamp}.json")
            
            # 执行详细比对
            self.log_message("开始详细比对文件...", "blue")
            
            # 收集所有差异
            all_differences = []
            
            # 比对每对文件
            for i in range(len(dfs)):
                for j in range(i+1, len(dfs)):
                    # 为每对文件创建独立的差异列表
                    file_pair_diffs = []
                    diff_count = 0
                    self.log_message(f"\n=== 开始比对: {file_names[i]} vs {file_names[j]} ===", "darkblue")
                    
                    # 检查行数差异
                    if len(dfs[i]) != len(dfs[j]):
                        self.log_message(f"⚠️ 行数差异: {file_names[i]}有{len(dfs[i])}行, {file_names[j]}有{len(dfs[j])}行", "orange")
                    
                    # 逐行比对
                    max_rows = min(len(dfs[i]), len(dfs[j]))
                    for row_idx in range(max_rows):
                        row_diff = False
                        diff_details = []
                        
                        # 逐列比对
                        for col in dfs[i].columns:
                            val1 = dfs[i].iloc[row_idx][col]
                            val2 = dfs[j].iloc[row_idx][col]
                            
                            if val1 != val2:
                                row_diff = True
                                diff_details.append(f"{col}: '{val1}' vs '{val2}'")
                        
                        if row_diff:
                            diff_count += 1
                            # 在UI中用不同颜色显示差异
                            self.log_message(f"🔴 行 {row_idx+1} 差异: {', '.join(diff_details)}", "red")
                            file_pair_diffs.append({
                                'row': row_idx+1,
                                'details': diff_details
                            })
                    
                    # 保存这对文件的差异
                    if file_pair_diffs:
                        all_differences.append({
                            'file_pair': f"{file_names[i]} vs {file_names[j]}",
                            'diffs': file_pair_diffs,
                            'total_diffs': diff_count
                        })
                    
                    # 输出文件比对摘要
                    summary_msg = f"📊 比对摘要: {file_names[i]} 和 {file_names[j]} - "
                    summary_msg += f"共发现 {diff_count} 处差异" if diff_count > 0 else "无差异"
                    self.log_message(summary_msg, "green" if diff_count == 0 else "orange")
            
            # 生成JSON格式报告
            report_data = {
                "report_time": timestamp,
                "compared_files": file_names,
                "total_differences": len(all_differences),
                "comparisons": []
            }
            
            for diff in all_differences:
                comparison = {
                    "file_pair": diff['file_pair'],
                    "total_differences": diff['total_diffs'],
                    "differences": []
                }
                
                for row_diff in diff['diffs']:
                    difference = {
                        "row": row_diff['row'],
                        "details": row_diff['details']
                    }
                    comparison['differences'].append(difference)
                
                report_data['comparisons'].append(comparison)
            
            # 保存JSON报告
            import json
            with open(report_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=4, ensure_ascii=False)
            
            self.log_message(f"\n比对完成! 详细报告已保存到: {report_path}", "blue")
            QMessageBox.information(self, "完成", "文件比对完成，报告已保存")
            
        except Exception as e:
            self.log_message(f"比对失败: {str(e)}", "red")
            QMessageBox.critical(self, "错误", f"比对失败: {str(e)}")

    def compare_with_database(self):
        """与数据库比对"""
        if not self.db_compare_file:
            QMessageBox.warning(self, "错误", "请先上传比对文件")
            return
            
        try:
            # 读取比对文件
            if self.db_compare_file.endswith('.xlsx'):
                compare_df = pd.read_excel(self.db_compare_file)
            else:
                compare_df = pd.read_csv(self.db_compare_file)
                
            # 检查表格格式
            if len(compare_df.columns) < 2:
                QMessageBox.warning(self, "错误", "比对文件必须包含至少2列数据")
                return
                
            # 读取数据库文件
            db_dirs = [d for d in os.listdir('data') if d in ('xlsx', 'csv')]
            db_files = []
            for dir_name in db_dirs:
                dir_path = os.path.join('data', dir_name)
                if os.path.isdir(dir_path):
                    files = [f for f in os.listdir(dir_path) if f.endswith(('.xlsx', '.csv'))]
                    if files:
                        db_files.extend([os.path.join(dir_name, f) for f in files])
            
            if not db_files:
                QMessageBox.warning(self, "错误", "数据库中没有文件")
                return
                
            # 获取最新的数据库文件
            db_files.sort(key=lambda x: os.path.getmtime(os.path.join('data', x)), reverse=True)
            db_path = os.path.join('data', db_files[0])
            if db_path.endswith('.xlsx'):
                db_df = pd.read_excel(db_path)
            else:
                db_df = pd.read_csv(db_path)
                
            # 检查数据库文件格式
            if len(db_df.columns) < 2:
                QMessageBox.warning(self, "错误", "数据库文件必须包含至少2列数据")
                return
                
            # 检查表头是否一致（忽略比对报告列）
            compare_cols = [col for col in compare_df.columns if not col.startswith('比对报告')]
            db_cols = [col for col in db_df.columns if not col.startswith('比对报告')]
            if compare_cols != db_cols:
                QMessageBox.warning(self, "错误", "比对文件与数据库文件格式不一致")
                return
                
            self.log_message("开始与数据库比对...")
            
            # 初始化统计信息
            total_items = 0
            new_items = 0
            changed_items = 0
            unchanged_items = 0
            
            # 创建Excel写入器
            writer = pd.ExcelWriter(db_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            db_df = pd.read_excel(db_path)
            
            # 创建带时间戳的报告列名
            report_col = f"比对报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # 保留所有历史比对报告列
            # 添加新的报告列
            db_df[report_col] = ''
                
            # 设置样式字典
            styles = {}
            
            # 获取需要比对的列（忽略所有以"比对报告"开头的列）
            compare_cols = [col for col in compare_df.columns 
                          if not col.startswith('比对报告')]
            
            # 逐行比对(包括第一行数据)
            for idx, row in compare_df.iterrows():
                # 跳过表头行(索引0)
                if idx == 0 and all(isinstance(val, str) for val in row.values):
                    continue
                    
                total_items += 1
                product_id = row.iloc[1]  # B列商品ID
                match = db_df[db_df.iloc[:, 1] == product_id]  # 在数据库B列查找匹配
                
                if match.empty:
                    # 新商品 - 添加到数据库
                    new_row = row.to_frame().T
                    new_row[report_col] = "新增商品: " + ", ".join([f"{col}: {row[col]}" for col in compare_df.columns if not col.startswith('比对报告')])
                    db_df = pd.concat([db_df, new_row], ignore_index=True)
                    # 标记新增商品的行索引
                    styles[len(db_df)-1] = 'new'  # 新增商品标记为'new'
                    new_items += 1
                    self.log_message(f"新增商品: ID {product_id}")
                else:
                    # 现有商品 - 比对数据
                    match_idx = match.index[0]
                    diff_cols = []
                    
                    for col in compare_df.columns:
                        if col != '比对报告' and row[col] != db_df.at[match_idx, col]:
                            diff_cols.append(col)
                    
                    if diff_cols:
                        # 有差异
                        report = "数据差异: "
                        report += ", ".join([f"{col}: {row[col]}→{db_df.at[match_idx, col]}" for col in diff_cols])
                        db_df.at[match_idx, report_col] = report
                        
                        # 更新数据
                        for col in diff_cols:
                            db_df.at[match_idx, col] = row[col]
                            
                        styles[match_idx] = 'changed'  # 标记为有差异
                        changed_items += 1
                        self.log_message(f"更新商品: ID {product_id} - 差异项: {', '.join(diff_cols)}")
                    else:
                        # 无差异
                        report = "数据一致: " + ", ".join([f"{col}: {row[col]}" for col in compare_df.columns if not col.startswith('比对报告')])
                        db_df.at[match_idx, report_col] = report
                        styles[match_idx] = 'unchanged'  # 标记为无差异
                        unchanged_items += 1
                        self.log_message(f"无差异商品: ID {product_id}")
            
            # 应用样式 - 使用更可靠的方式
            from openpyxl.styles import PatternFill
            from openpyxl import load_workbook
            
            # 保存数据并应用样式 - 使用更可靠的方式
            try:
                # 创建临时文件路径
                temp_path = db_path + '.tmp'
                
                # 使用openpyxl直接创建工作簿
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                
                # 写入表头
                for col_num, col_name in enumerate(db_df.columns, 1):
                    ws.cell(row=1, column=col_num, value=col_name)
                
                # 写入数据
                for row_num, row in enumerate(db_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
                
                # 定义三种颜色填充
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 差异-红色
                green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 无差异-绿色
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # 新增-黄色
                
                # 获取报告列索引(1-based)
                report_col_idx = db_df.columns.get_loc(report_col) + 1
                
                # 应用颜色到报告列和整行
                for idx, status in styles.items():
                    row_idx = idx + 2  # Excel行索引从1开始，且跳过表头
                    
                    # 确保样式字典中的状态被正确识别
                    if status == 'new':  # 新增商品 - 黄色
                        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        ws.cell(row=row_idx, column=report_col_idx).fill = fill
                        self.log_message(f"应用黄色到新增商品行{row_idx}的报告列")
                    elif status == 'changed':  # 有差异 - 红色
                        fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        ws.cell(row=row_idx, column=report_col_idx).fill = fill
                    else:  # 无差异 - 绿色
                        fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                        ws.cell(row=row_idx, column=report_col_idx).fill = fill
                
                # 保存到临时文件
                wb.save(temp_path)
                
                # 验证临时文件
                test_wb = load_workbook(temp_path)
                test_wb.close()
                
                # 替换原文件
                if os.path.exists(db_path):
                    os.remove(db_path)
                os.rename(temp_path, db_path)
                
                self.log_message("数据保存成功")
            except Exception as e:
                self.log_message(f"数据保存失败: {str(e)}", "red")
                # 恢复备份文件
                backup_path = db_path + ".bak"
                if os.path.exists(backup_path):
                    shutil.copy2(backup_path, db_path)
                else:
                    self.log_message("无备份文件可恢复", "red")
            
            # 输出总结报告
            summary = f"\n比对总结报告:\n"
            summary += f"比对时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            summary += f"总比对商品数: {total_items}\n"
            summary += f"新增商品数: {new_items}\n"
            summary += f"有差异商品数: {changed_items}\n"
            summary += f"无差异商品数: {unchanged_items}\n"
            
            self.info_display.append(summary)
            QMessageBox.information(self, "完成", "数据库比对完成")
            
        except Exception as e:
            self.log_message(f"数据库比对失败: {str(e)}", "red")
            QMessageBox.critical(self, "错误", f"数据库比对失败: {str(e)}")
